# 5x5 grid letters, missing I
alphabet = 'ABCDEFGHJKLMNOPQRSTUVWXYZ'

def grid2xy(false_easting, false_northing, gridsizes, grid_ref):
    '''Convert grid reference to coordinates'''
    # false easting and northing
    easting = -false_easting
    northing = -false_northing

    # convert letter(s) to easting and northing offset
    for n in range(0, len(gridsizes)):
        letter = grid_ref[n]
        idx = alphabet.index(letter)
        col = (idx % 5)
        row = 4 - ((idx) / 5)
        easting += (col * gridsizes[n])
        northing += (row * gridsizes[n])

    # numeric components of grid reference
    grid_ref = grid_ref[len(gridsizes):] # remove the letters
    e = '{:0<5}'.format(grid_ref[0:len(grid_ref)/2])
    e = '{}.{}'.format(e[0:5],e[5:])
    n = '{:0<5}'.format(grid_ref[len(grid_ref)/2:])
    n = '{}.{}'.format(n[0:5],n[5:])
    easting += float(e)
    northing += float(n)

    return easting, northing

def british2xy(grid_ref):
    false_easting = 1000000
    false_northing = 500000
    gridsizes = [500000, 100000]
    return grid2xy(false_easting, false_northing, gridsizes, grid_ref)

def irish2xy(grid_ref):
    false_easting = 0
    false_northing = 0
    gridsizes = [100000]
    return grid2xy(false_easting, false_northing, gridsizes, grid_ref)

grid_ref = 'NO19001640' #ABNTY
easting, northing = british2xy(grid_ref)
print(easting, northing)

grid_ref = 'W70308320' #BABOR
easting, northing = irish2xy(grid_ref)
print(easting, northing)

# PROJ.4 projection definitions
#import pyproj
#crs_british = pyproj.Proj(init='EPSG:27700')
#crs_irish = pyproj.Proj(init='EPSG:29903')
#crs_wgs84 = pyproj.Proj(init='EPSG:4326')

#lng, lat = pyproj.transform(crs_british, crs_wgs84, easting, northing)
#lng, lat = pyproj.transform(crs_irish, crs_wgs84, easting, northing)

'''
# prepare the output shapefile
import fiona
from collections import OrderedDict
crs = {'no_defs': True, 'ellps': 'WGS84', 'datum': 'WGS84', 'proj': 'longlat'}
properties = OrderedDict([
    ('surveyid', 'int'),
    ('siteid', 'int'),
    ('gridref', 'str'),
    ('date', 'str'),
    ('river', 'str'),
])
layer = fiona.open('rhs.shp', 'w', crs=crs, driver='ESRI Shapefile', schema={'geometry': 'Point', 'properties': properties})

# open the spreadsheet
from openpyxl import load_workbook
wb = load_workbook(filename='River_Habitat_Survey-Details_and_Scores.xlsx', use_iterators=True)
ws = wb.get_sheet_by_name('Sheet1')
rows_iter = ws.iter_rows()
header_row = rows_iter.next()
# iterate over the records
for row in rows_iter:
    grid_ref = row[2].value
    if grid_ref[1] in alphabet:
        # grid reference is for the british national grid
        easting, northing = british2xy(grid_ref)
        lng, lat = pyproj.transform(crs_british, crs_wgs84, easting, northing)
    else:
        # grid reference is for the irish grid
        easting, northing = irish2xy(grid_ref)
        lng, lat = pyproj.transform(crs_irish, crs_wgs84, easting, northing)

    # write output feature
    properties = OrderedDict([
        ('surveyid', row[0].value),
        ('siteid', row[1].value),
        ('gridref', grid_ref),
        ('date', str(row[3].value)),
        ('river', row[4].value),
    ])
    feature = {'geometry': {'type': 'Point', 'coordinates': (lng, lat)}, 'properties': properties}
    layer.write(feature)
layer.close()
'''
